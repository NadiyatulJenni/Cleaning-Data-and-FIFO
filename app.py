import streamlit as st
import pandas as pd
import io
import os
import re
import datetime
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# --- KONFIGURASI HALAMAN ---
st.set_page_config(page_title="FIFO Master", layout="wide")

# Fungsi format angka Indonesia (Ribuan: titik, Desimal: koma)
def format_indonesia(value, fmt_type):
    if pd.isna(value) or value == "": return ""
    try:
        val = float(value)
        if fmt_type == "integers":
            return "{:,.0f}".format(val).replace(",", ".")
        elif fmt_type == "decimal number" or fmt_type == "Currency (Rp)":
            res = "{:,.2f}".format(val)
            parts = res.split(".")
            parts[0] = parts[0].replace(",", ".")
            return ",".join(parts)
        return value
    except: return value

st.title("📦 Streamline FIFO by Nadd")
st.markdown("Aplikasi Cleaning Data & Integrasi FIFO - By Nadiyatul Jenni")

# Sidebar Upload
st.sidebar.header("📁 Upload Data Sumber")
file_so = st.sidebar.file_uploader("SO / Stok Awal", type=['xlsx'])
file_masuk = st.sidebar.file_uploader("Barang Masuk", type=['xlsx'])
file_keluar = st.sidebar.file_uploader("Barang Keluar", type=['xlsx'])

if 'cleaned_data' not in st.session_state:
    st.session_state['cleaned_data'] = {}

tab_cleaning, tab_fifo = st.tabs(["🧹 Data Cleaning", "🔄 Integrasi FIFO"])

# ==========================================
# TAB 1: CLEANING DATA (LOGIKA DIPERKUAT)
# ==========================================
with tab_cleaning:
    files = {"SO": file_so, "Masuk": file_masuk, "Keluar": file_keluar}
    active_files = {k: v for k, v in files.items() if v is not None}
    
    if not active_files:
        st.info("Silakan upload file Excel di sidebar untuk memulai.")
    else:
        pilihan = st.selectbox("Pilih file yang akan dikelola:", list(active_files.keys()))
        curr_file = active_files[pilihan]
        
        col_c1, col_c2, col_c3 = st.columns(3)
        with col_c1:
            h_row = st.number_input(f"Baris Header {pilihan}:", min_value=1, value=1)
        with col_c2:
            d_start = st.number_input(f"Data Mulai Baris:", min_value=1, value=h_row+1)
        with col_c3:
            f_word = st.text_input("Hapus baris mengandung kata (Contoh: TOTAL):")

        do_trim = st.checkbox("Gunakan Fungsi TRIM", value=True)

        df_raw = pd.read_excel(curr_file, header=h_row-1)
        skip_n = d_start - (h_row + 1)
        if skip_n > 0: df_raw = df_raw.iloc[skip_n:].reset_index(drop=True)
        
        if f_word:
            mask = df_raw.apply(lambda r: r.astype(str).str.contains(f_word, case=False).any(), axis=1)
            df_raw = df_raw[~mask].reset_index(drop=True)

        st.subheader("Preview Data Asli")
        st.dataframe(df_raw.head(10))

        st.divider()
        st.subheader("⚙️ Pengaturan Kolom & Format")
        
        selected_info = []
        for col in df_raw.columns:
            if "Unnamed" in str(col): continue
            c1, c2, c3 = st.columns([1, 2, 2])
            with c1:
                is_keep = st.checkbox("Keep", value=True, key=f"c_{pilihan}_{col}")
            with c2:
                new_n = st.text_input("Rename", value=str(col), key=f"r_{pilihan}_{col}")
            with c3:
                f_type = st.selectbox("Format", 
                    ["Text", "Date", "Date Time", "Currency (Rp)", "integers", "decimal number", "Percent (%)"], 
                    key=f"f_{pilihan}_{col}")
            if is_keep:
                selected_info.append({"old": col, "new": new_n, "type": f_type})

        if st.button(f"🔥 Simpan Cleaning {pilihan}"):
            with st.spinner("Memproses seluruh data..."):
                cols = [i['old'] for i in selected_info]
                names = {i['old']: i['new'] for i in selected_info}
                df_clean = df_raw[cols].rename(columns=names).copy()

                if do_trim:
                    df_clean = df_clean.applymap(lambda x: x.strip() if isinstance(x, str) else x)

                for i in selected_info:
                    target = i['new']
                    ft = i['type']
                    
                    if ft == "Date":
                        df_clean[target] = pd.to_datetime(df_clean[target], errors='coerce').dt.date
                    elif ft == "Date Time":
                        df_clean[target] = pd.to_datetime(df_clean[target], errors='coerce')
                    elif ft in ["integers", "decimal number", "Currency (Rp)"]:
                        # Deep Clean: Menghapus titik ribuan dan Rp sebelum konversi agar tidak jadi 0
                        def clean_numeric_strict(val):
                            if pd.isna(val) or val == "": return 0.0
                            s = str(val).replace('Rp', '').replace(' ', '')
                            if ',' in s and '.' in s: s = s.replace('.', '').replace(',', '.')
                            elif ',' in s: s = s.replace(',', '.')
                            s = re.sub(r'[^\d.]', '', s)
                            return pd.to_numeric(s, errors='coerce') if s != "" else 0.0
                        
                        df_clean[target] = df_clean[target].apply(clean_numeric_strict).astype(float)

                st.session_state['cleaned_data'][pilihan] = df_clean
                st.success(f"Berhasil! {len(df_clean)} baris data {pilihan} disimpan.")

                # Format for display
                df_display_clean = df_clean.copy()
                # Format currency
                currency_cols_clean = [col for col in df_display_clean.columns if any(x in col for x in ['Harga', 'Total', 'Nilai'])]
                for col in currency_cols_clean:
                    df_display_clean[col] = pd.to_numeric(df_display_clean[col], errors='coerce').apply(lambda x: f"Rp {x:,.2f}" if pd.notna(x) and x != 0 else ("Rp 0.00" if x == 0 else ""))
                # Format date
                date_cols_clean = [col for col in df_display_clean.columns if 'Tanggal' in col or 'Date' in col]
                for col in date_cols_clean:
                    if pd.api.types.is_datetime64_any_dtype(df_display_clean[col]):
                        df_display_clean[col] = df_display_clean[col].dt.strftime('%d/%m/%Y')
                    elif pd.api.types.is_datetime64_any_dtype(df_display_clean[col].dt):
                        df_display_clean[col] = df_display_clean[col].dt.strftime('%d/%m/%Y %H:%M:%S')
                # Format integers
                int_cols_clean = [col for col in df_display_clean.columns if any(x in col for x in ['Qty', 'Stok'])]
                for col in int_cols_clean:
                    df_display_clean[col] = pd.to_numeric(df_display_clean[col], errors='coerce').apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "")

                st.dataframe(df_display_clean.head(10))

                # Download cleaned data as Excel
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df_clean.to_excel(writer, index=False, sheet_name='Cleaned_Data')
                    ws = writer.sheets['Cleaned_Data']
                    for col_idx, col_name in enumerate(df_clean.columns, 1):
                        let = get_column_letter(col_idx)
                        ws[f'{let}1'].font = Font(bold=True)
                        ws.column_dimensions[let].width = 20
                        for row in range(2, len(df_clean) + 2):
                            cell = ws[f'{let}{row}']
                            if pd.isna(cell.value): continue
                            # Apply formats based on type
                            for i in selected_info:
                                if i['new'] == col_name:
                                    ft = i['type']
                                    if ft == "Date":
                                        cell.number_format = 'dd/mm/yyyy'
                                    elif ft == "Date Time":
                                        cell.number_format = 'dd/mm/yyyy hh:mm:ss'
                                    elif ft == "Currency (Rp)":
                                        cell.number_format = '"Rp" #,##0.00'
                                    elif ft == "integers":
                                        cell.number_format = '#,##0'
                                    elif ft == "decimal number":
                                        cell.number_format = '#,##0.00'
                                    break

                buffer.seek(0)
                st.download_button(
                    label="📥 Download Cleaned Data (Excel)",
                    data=buffer,
                    file_name=f"cleaned_{pilihan.lower()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

# ==========================================
# TAB 2: INTEGRASI FIFO (REVISI FIX)
# ==========================================
with tab_fifo:
    st.header("🔄 Proses Logika FIFO Terintegrasi")
    
    df_so_raw = st.session_state['cleaned_data'].get("SO")
    df_in_raw = st.session_state['cleaned_data'].get("Masuk")
    df_out_raw = st.session_state['cleaned_data'].get("Keluar")

    siap_proses = (df_in_raw is not None or df_so_raw is not None) and df_out_raw is not None

    if not siap_proses:
        st.warning("⚠️ Minimal harus ada (SO + Keluar) atau (Masuk + Keluar) untuk memulai.")
    else:
        d_so_ui = df_so_raw if df_so_raw is not None else pd.DataFrame(columns=['-'])
        d_in_ui = df_in_raw if df_in_raw is not None else pd.DataFrame(columns=['-'])
        d_out_ui = df_out_raw

        st.subheader("🛠️ Mapping Kolom (3 Tabel Mandiri)")
        m1, m2, m3 = st.columns(3)
        with m1:
            st.info("📌 Kolom SO (Stok Awal)")
            map_p_so = st.selectbox("Produk (SO)", ['-'] + list(d_so_ui.columns), key="m_p_so")
            map_q_so = st.selectbox("Qty (SO)", ['-'] + list(d_so_ui.columns), key="m_q_so")
            map_h_so = st.selectbox("Harga Satuan (SO)", ['-'] + list(d_so_ui.columns), key="m_h_so")
        with m2:
            st.success("📩 Kolom Barang Masuk")
            map_t_in = st.selectbox("Tanggal (Masuk)", ['-'] + list(d_in_ui.columns), key="m_t_in")
            map_p_in = st.selectbox("Produk (Masuk)", ['-'] + list(d_in_ui.columns), key="m_p_in")
            map_q_in = st.selectbox("Qty (Masuk)", ['-'] + list(d_in_ui.columns), key="m_q_in")
            map_h_in = st.selectbox("Harga Satuan (Masuk)", ['-'] + list(d_in_ui.columns), key="m_h_in")
        with m3:
            st.error("📤 Kolom Barang Keluar")
            map_t_out = st.selectbox("Tanggal (Keluar)", ['-'] + list(d_out_ui.columns), key="m_t_out")
            map_p_out = st.selectbox("Produk (Keluar)", ['-'] + list(d_out_ui.columns), key="m_p_out")
            map_q_out = st.selectbox("Qty (Keluar)", ['-'] + list(d_out_ui.columns), key="m_q_out")

        st.subheader("➕ Kolom Tambahan (Optional)")
        if 'opt_configs' not in st.session_state:
            st.session_state.opt_configs = [{'label': 'Info_1', 'so': '-', 'in': '-', 'out': '-'}]

        for i, config in enumerate(st.session_state.opt_configs):
            c_o1, c_o2, c_o3, c_o4, c_o5 = st.columns([2, 2, 2, 2, 1])
            with c_o1: config['label'] = st.text_input(f"Nama Kolom Result {i+1}", config['label'], key=f"lo_{i}")
            with c_o2: config['so'] = st.selectbox(f"Ambil dari SO", ['-'] + list(d_so_ui.columns), key=f"so_o_{i}")
            with c_o3: config['in'] = st.selectbox(f"Ambil dari Masuk", ['-'] + list(d_in_ui.columns), key=f"in_o_{i}")
            with c_o4: config['out'] = st.selectbox(f"Ambil dari Keluar", ['-'] + list(d_out_ui.columns), key=f"out_o_{i}")
            with c_o5: 
                if st.button("❌", key=f"remove_{i}"):
                    st.session_state.opt_configs.pop(i)
                    st.rerun()

        if st.button("➕ Tambah Baris Optional"):
            st.session_state.opt_configs.append({'label': f'Info_{len(st.session_state.opt_configs)+1}', 'so': '-', 'in': '-', 'out': '-'})
            st.rerun()

        opt_configs = st.session_state.opt_configs

        if st.button("🚀 JALANKAN CORE FIFO"):
            # Validation: Ensure mandatory mappings are not '-'
            errors = []
            if df_so_raw is not None:
                if map_p_so == '-': errors.append("Produk (SO) tidak boleh '-'")
                if map_q_so == '-': errors.append("Qty (SO) tidak boleh '-'")
                if map_h_so == '-': errors.append("Harga Satuan (SO) tidak boleh '-'")
            if df_in_raw is not None:
                if map_t_in == '-': errors.append("Tanggal (Masuk) tidak boleh '-'")
                if map_p_in == '-': errors.append("Produk (Masuk) tidak boleh '-'")
                if map_q_in == '-': errors.append("Qty (Masuk) tidak boleh '-'")
                if map_h_in == '-': errors.append("Harga Satuan (Masuk) tidak boleh '-'")
            if map_t_out == '-': errors.append("Tanggal (Keluar) tidak boleh '-'")
            if map_p_out == '-': errors.append("Produk (Keluar) tidak boleh '-'")
            if map_q_out == '-': errors.append("Qty (Keluar) tidak boleh '-'")

            if errors:
                st.error("Mapping kolom wajib tidak boleh '-'. Silakan pilih kolom yang valid.")
                for err in errors:
                    st.error(f"- {err}")
            else:
                try:
                    # Proses Data Masuk & SO
                    if df_in_raw is not None:
                        d_in = df_in_raw.copy()
                        d_in['TGL_FIFO'] = pd.to_datetime(d_in[map_t_in])
                        d_in['QTY_FIFO'] = pd.to_numeric(d_in[map_q_in]).fillna(0)
                        d_in['PRC_FIFO'] = pd.to_numeric(d_in[map_h_in]).fillna(0)
                        d_in['SOURCE'] = 1
                    else: d_in = pd.DataFrame()

                    if df_so_raw is not None:
                        d_so = df_so_raw.copy()
                        ref_y = d_in['TGL_FIFO'].dt.year.min() if not d_in.empty else pd.Timestamp.now().year
                        d_so['TGL_FIFO'] = pd.Timestamp(year=int(ref_y)-1, month=12, day=31)
                        d_so['QTY_FIFO'] = pd.to_numeric(d_so[map_q_so]).fillna(0)
                        d_so['PRC_FIFO'] = pd.to_numeric(d_so[map_h_so]).fillna(0)
                        d_so['SOURCE'] = 0
                    else: d_so = pd.DataFrame()

                    # Merge Inventory
                    inv_merged = pd.concat([d_so, d_in], ignore_index=True).sort_values(['TGL_FIFO', 'SOURCE'])

                    # Proses Data Keluar
                    d_out = df_out_raw.copy()
                    d_out['TGL_FIFO'] = pd.to_datetime(d_out[map_t_out])
                    d_out['QTY_FIFO'] = pd.to_numeric(d_out[map_q_out]).fillna(0)
                    d_out = d_out.sort_values(by=['TGL_FIFO']).reset_index(drop=True)

                    hasil = []
                    prod_keys = set(inv_merged[map_p_in if not d_in.empty else map_p_so].unique()).union(set(d_out[map_p_out].unique()))

                    for kode in prod_keys:
                        if pd.isna(kode) or kode == "-": continue

                        stok_antrean = []
                        masuk_produk = inv_merged[inv_merged[map_p_in if not d_in.empty else map_p_so] == kode].to_dict('records')
                        keluar_produk = d_out[d_out[map_p_out] == kode].to_dict('records')

                        # Proses Barang Masuk (SO + Masuk)
                        for m in masuk_produk:
                            opt = {}
                            for oc in opt_configs:
                                col = oc['so'] if m['SOURCE'] == 0 else oc['in']
                                if col == "-":
                                    opt[oc['label']] = pd.NA
                                else:
                                    opt[oc['label']] = m.get(col, pd.NA)
                            stok_antrean.append({
                                'qty': m['QTY_FIFO'],
                                'prc': m['PRC_FIFO'],
                                'opt': opt
                            })

                            hasil.append({
                                'Tanggal': m['TGL_FIFO'],
                                'Produk': kode,
                                **opt,
                                'Qty Masuk': m['QTY_FIFO'],
                                'Harga Satuan Masuk': m['PRC_FIFO'],
                                'Total Masuk': m['QTY_FIFO'] * m['PRC_FIFO'],
                                'Qty Keluar': 0,
                                'Harga Satuan Keluar': 0,
                                'Total Keluar': 0,
                                'Keterangan': 'BARANG MASUK'
                            })

                        # Proses Barang Keluar
                        for k in keluar_produk:
                            qty_perlu = k['QTY_FIFO']
                            while qty_perlu > 0:
                                if not stok_antrean:
                                    opt_empty = {}
                                    for oc in opt_configs:
                                        col = oc['out']
                                        if col == "-":
                                            opt_empty[oc['label']] = pd.NA
                                        else:
                                            opt_empty[oc['label']] = k.get(col, pd.NA)
                                    hasil.append({
                                        'Tanggal': k['TGL_FIFO'], 'Produk': kode, **opt_empty,
                                        'Qty Masuk': 0, 'Harga Satuan Masuk': 0, 'Total Masuk': 0,
                                        'Qty Keluar': qty_perlu, 'Harga Satuan Keluar': 0, 'Total Keluar': 0,
                                        'Keterangan': 'OUT OF STOCK'
                                    })
                                    qty_perlu = 0
                                else:
                                    stok_skrg = stok_antrean[0]
                                    if stok_skrg['qty'] <= qty_perlu:
                                        ambil = stok_skrg['qty']; qty_perlu -= ambil; hna = stok_skrg['prc']
                                        opt_ambil = stok_skrg['opt']
                                        stok_antrean.pop(0)
                                    else:
                                        ambil = qty_perlu; qty_perlu = 0; hna = stok_skrg['prc']
                                        opt_ambil = stok_skrg['opt']
                                        stok_skrg['qty'] -= ambil

                                    # For BARANG KELUAR, use opt from keluar table
                                    opt_keluar = {}
                                    for oc in opt_configs:
                                        col = oc['out']
                                        if col == "-":
                                            opt_keluar[oc['label']] = pd.NA
                                        else:
                                            opt_keluar[oc['label']] = k.get(col, pd.NA)

                                    hasil.append({
                                        'Tanggal': k['TGL_FIFO'], 'Produk': kode, **opt_keluar,
                                        'Qty Masuk': 0, 'Harga Satuan Masuk': 0, 'Total Masuk': 0,
                                        'Qty Keluar': ambil, 'Harga Satuan Keluar': hna, 'Total Keluar': ambil * hna,
                                        'Keterangan': 'BARANG KELUAR'
                                    })

                    df_final = pd.DataFrame(hasil)
                    if not df_final.empty:
                        df_final = df_final.sort_values(by=['Produk', 'Tanggal']).reset_index(drop=True)
                        df_final['Stok'] = df_final.groupby('Produk')['Qty Masuk'].cumsum() - df_final.groupby('Produk')['Qty Keluar'].cumsum()
                        df_final['Nilai Stok'] = df_final.groupby('Produk')['Total Masuk'].cumsum() - df_final.groupby('Produk')['Total Keluar'].cumsum()

                        # Update Keterangan for last row per product
                        df_final['is_last'] = df_final.groupby('Produk').cumcount(ascending=False) == 0
                        def assign_final_keterangan(row):
                            if row['is_last']:
                                if row['Keterangan'] == 'OUT OF STOCK':
                                    return 'STOK SEKARANG (OUT OF STOCK)'
                                else:
                                    return 'STOK SEKARANG'
                            return row['Keterangan']
                        df_final['Keterangan'] = df_final.apply(assign_final_keterangan, axis=1)
                        df_final = df_final.drop(columns=['is_last'])

                        st.success("✅ BERHASIL")

                        # Format for display
                        df_display = df_final.copy()
                        # Format currency
                        currency_cols = [col for col in df_display.columns if any(x in col for x in ['Harga', 'Total', 'Nilai'])]
                        for col in currency_cols:
                            df_display[col] = pd.to_numeric(df_display[col], errors='coerce').apply(lambda x: f"Rp {x:,.2f}" if pd.notna(x) and x != 0 else ("Rp 0.00" if x == 0 else ""))
                        # Format date
                        if 'Tanggal' in df_display.columns:
                            df_display['Tanggal'] = df_display['Tanggal'].dt.strftime('%d/%m/%Y %H:%M:%S')
                        # Format integers
                        int_cols = [col for col in df_display.columns if any(x in col for x in ['Qty', 'Stok'])]
                        for col in int_cols:
                            df_display[col] = pd.to_numeric(df_display[col], errors='coerce').apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "")

                        st.dataframe(df_display)

                        # Download FIFO result as Excel
                        buffer_fifo = io.BytesIO()
                        with pd.ExcelWriter(buffer_fifo, engine='openpyxl') as writer:
                            df_final.to_excel(writer, index=False, sheet_name='FIFO_Result')
                            ws = writer.sheets['FIFO_Result']
                            for col_idx, col_name in enumerate(df_final.columns, 1):
                                let = get_column_letter(col_idx)
                                ws[f'{let}1'].font = Font(bold=True)
                                ws.column_dimensions[let].width = 20
                                for row in range(2, len(df_final) + 2):
                                    cell = ws[f'{let}{row}']
                                    if pd.isna(cell.value): continue
                                    # Apply formats
                                    if col_name == 'Tanggal':
                                        cell.number_format = 'dd/mm/yyyy hh:mm:ss'
                                    elif any(x in col_name for x in ['Harga', 'Total', 'Nilai']):
                                        cell.number_format = '"Rp" #,##0.00'
                                    elif any(x in col_name for x in ['Qty', 'Stok']):
                                        cell.number_format = '#,##0'

                        buffer_fifo.seek(0)
                        st.download_button(
                            label="📥 Download FIFO Result (Excel)",
                            data=buffer_fifo,
                            file_name="fifo_result.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                except Exception as e:
                    st.error(f"Error: {e}")

# Footer watermark
st.markdown("---")
st.markdown(
    """
    <div style="text-align: center; font-size: small; color: gray;">
        © 2026 | Developed by <a href="https://www.linkedin.com/in/nadiyatul-jenni" target="_blank">Nadiyatul Jenni</a>
    </div>
    """,
    unsafe_allow_html=True
)
